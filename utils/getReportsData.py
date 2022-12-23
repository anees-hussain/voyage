import os
import pyautogui as pag
from time import sleep
from openpyxl import load_workbook
from datetime import datetime

dt = datetime.now()

mediapath = "D:/_projects/python/voyage/media"
primarySales = "C:/Users/HP/OneDrive/Desktop/voyage-reports/primary-sales.xlsx"
secondarySales = "C:/Users/HP/OneDrive/Desktop/voyage-reports/secondary-sales.xlsx"
stockReport = "C:/Users/HP/OneDrive/Desktop/voyage-reports/stock-report.xlsx"

def iterate_column(workbook_path, column_index):
    # Opens the workbook and updates it to latest version
    os.system(f'cmd /c "start {workbook_path}"')

    sleep(3)
    save = pag.locateCenterOnScreen(f"{mediapath}/save-icon2.png", confidence=0.9)
    pag.moveTo(save.x, save.y, 0.5)
    pag.click()

    sleep(2)
    close = pag.locateCenterOnScreen(f"{mediapath}/close-button.png", confidence=0.9)
    pag.moveTo(close.x, close.y, 0.5)
    pag.click()

    # Loads workbook to extract data
    wb = load_workbook(workbook_path)
    ws = wb.active
    target_column = ws.iter_cols(column_index, column_index, values_only=True)
    totalValue = 0
    for value in target_column:
        sum = 0
        for v in range(len(value)):
            if type(value[v]) == int or type(value[v]) == float:
                sum = sum + value[v]
            else:
                pass
        totalValue = round(sum)
    wb.close()
    return totalValue    

def getPrimarySales():
    totalPrimarySale = iterate_column(workbook_path=primarySales, column_index=11)
    print = f"Primary Lifting UC: {totalPrimarySale}"
    return print

def getSecondarySales():
    totalSecondarySale = iterate_column(workbook_path=secondarySales, column_index=11)
    print = f"Secondary Sale UC: {totalSecondarySale}"
    return print

def getStockData():
    totalPhyStock = iterate_column(workbook_path=stockReport, column_index=7)
    totalUcStock = iterate_column(workbook_path=stockReport, column_index=9)
    totalStockAmount = iterate_column(workbook_path=stockReport, column_index=13)
    stockDays = round((totalUcStock / 16400) * 26)
    
    print = f"""Total Stock (PHC): {totalPhyStock}
Total Stock (UC): {totalUcStock}
Total Stock Amount: {totalStockAmount}
Stock Days: ({stockDays})"""
    return print

def getData():
    primary = getPrimarySales()
    secondary = getSecondarySales()
    stock = getStockData()
    return f'Date: {dt.date().strftime("%d-%m-%Y")}\n{primary}\n{secondary}\n{stock}'